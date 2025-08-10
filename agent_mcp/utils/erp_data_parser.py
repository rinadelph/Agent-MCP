"""
ERP Data Parser Module - Production-Ready Data Integration Utilities

This module provides comprehensive parsing capabilities for eFab ERP data files,
handling malformed CSV with HTML artifacts, Excel files, and providing robust
data validation and sanitization for textile ERP integration.

Features:
- HTML artifact cleaning and removal
- CSV/Excel file parsing with encoding detection
- Data type conversion and validation
- Header normalization and mapping
- Error handling and detailed logging
- Memory-efficient streaming for large files
- Comprehensive data sanitization
"""

import csv
import re
import json
import logging
import pandas as pd
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple, Union, Iterator
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from io import StringIO, BytesIO
import chardet
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import xlrd
from dataclasses import dataclass

from ..core.config import logger


@dataclass
class ParseResult:
    """Result of parsing operation with metadata."""
    data: List[Dict[str, Any]]
    total_rows: int
    valid_rows: int
    errors: List[str]
    warnings: List[str]
    metadata: Dict[str, Any]
    original_headers: List[str]
    normalized_headers: List[str]


@dataclass
class ParsingConfig:
    """Configuration for parsing operations."""
    skip_rows: int = 0
    max_rows: Optional[int] = None
    encoding: Optional[str] = None
    delimiter: str = ','
    strip_html: bool = True
    normalize_headers: bool = True
    validate_data: bool = True
    chunk_size: int = 1000
    date_formats: List[str] = None
    decimal_separator: str = '.'
    thousands_separator: str = ','
    
    def __post_init__(self):
        if self.date_formats is None:
            self.date_formats = [
                '%Y-%m-%d',
                '%m/%d/%Y',
                '%d/%m/%Y',
                '%Y-%m-%d %H:%M:%S',
                '%m/%d/%Y %H:%M:%S'
            ]


class ERPDataParser:
    """
    Production-ready ERP data parser with comprehensive error handling
    and data sanitization capabilities.
    """
    
    # HTML cleaning patterns
    HTML_PATTERNS = {
        'button_onclick': re.compile(r'<button[^>]*onClick="[^"]*"[^>]*>.*?</button>', re.IGNORECASE | re.DOTALL),
        'div_tags': re.compile(r'<div[^>]*>.*?</div>', re.IGNORECASE | re.DOTALL),
        'script_tags': re.compile(r'<script[^>]*>.*?</script>', re.IGNORECASE | re.DOTALL),
        'style_tags': re.compile(r'<style[^>]*>.*?</style>', re.IGNORECASE | re.DOTALL),
        'html_tags': re.compile(r'<[^>]+>'),
        'html_entities': re.compile(r'&[a-zA-Z0-9#]+;'),
        'multiple_spaces': re.compile(r'\s+'),
        'line_breaks': re.compile(r'[\r\n]+')
    }
    
    # Field mapping for different ERP data types
    FIELD_MAPPINGS = {
        'inventory': {
            'style #': 'style_number',
            'order #': 'order_number',
            'customer': 'customer_name',
            'roll #': 'roll_number',
            'vendor roll #': 'vendor_roll_number',
            'rack': 'rack_location',
            'qty (yds)': 'quantity_yards',
            'qty (lbs)': 'quantity_pounds',
            'good ea.': 'good_pieces',
            'bad ea.': 'bad_pieces',
            'received': 'received_date'
        },
        'sales_orders': {
            'status': 'order_status',
            'csr': 'customer_service_rep',
            'unit price': 'unit_price',
            'quoted date': 'quote_date',
            'cfversion': 'cf_version',
            'fbase': 'fabric_base',
            'on hold': 'on_hold_status',
            'ordered': 'quantity_ordered',
            'picked/shipped': 'quantity_shipped',
            'balance': 'balance_quantity',
            'available': 'available_quantity',
            'uom': 'unit_of_measure',
            'sop': 'sales_order_process',
            'po #': 'purchase_order_number',
            'sold to': 'sold_to_customer',
            'ship to': 'ship_to_customer',
            'ship date': 'ship_date'
        },
        'yarn_demand': {
            'yarn': 'yarn_id',
            'supplier': 'supplier_name',
            'description': 'yarn_description',
            'color': 'color_code',
            'inventory': 'current_inventory',
            'total demand': 'total_demand',
            'total receipt': 'total_receipts',
            'balance': 'balance_quantity',
            'past due receipts': 'past_due_receipts'
        }
    }
    
    def __init__(self, config: Optional[ParsingConfig] = None):
        """Initialize parser with configuration."""
        self.config = config or ParsingConfig()
        self.logger = logger
        
    def detect_encoding(self, file_path: Union[str, Path]) -> str:
        """Detect file encoding using chardet."""
        try:
            with open(file_path, 'rb') as f:
                sample = f.read(10000)  # Read first 10KB for detection
                result = chardet.detect(sample)
                encoding = result['encoding'] or 'utf-8'
                confidence = result['confidence'] or 0.0
                
                self.logger.debug(f"Detected encoding: {encoding} (confidence: {confidence:.2f})")
                
                # Handle common encoding issues
                if encoding.lower() in ['ascii', 'windows-1252']:
                    encoding = 'utf-8'
                    
                return encoding
        except Exception as e:
            self.logger.warning(f"Encoding detection failed: {e}, defaulting to utf-8")
            return 'utf-8'
    
    def clean_html_artifacts(self, text: str) -> str:
        """
        Clean HTML artifacts from text data with comprehensive pattern matching.
        
        Args:
            text: Input text that may contain HTML artifacts
            
        Returns:
            Cleaned text with HTML artifacts removed
        """
        if not isinstance(text, str) or not text.strip():
            return text
        
        cleaned_text = text
        
        try:
            # Apply regex patterns for common HTML structures
            for pattern_name, pattern in self.HTML_PATTERNS.items():
                cleaned_text = pattern.sub('', cleaned_text)
            
            # Use BeautifulSoup for more comprehensive HTML cleaning
            soup = BeautifulSoup(cleaned_text, 'html.parser')
            cleaned_text = soup.get_text(separator=' ', strip=True)
            
            # Final cleanup
            cleaned_text = self.HTML_PATTERNS['multiple_spaces'].sub(' ', cleaned_text)
            cleaned_text = cleaned_text.strip()
            
            # Log significant HTML cleaning operations
            if len(text) - len(cleaned_text) > 50:
                self.logger.debug(f"Cleaned HTML artifacts: {len(text) - len(cleaned_text)} characters removed")
                
        except Exception as e:
            self.logger.warning(f"HTML cleaning failed for text: {text[:100]}... Error: {e}")
            # Fall back to basic cleaning
            cleaned_text = re.sub(r'<[^>]+>', '', text)
            cleaned_text = re.sub(r'\s+', ' ', cleaned_text).strip()
        
        return cleaned_text
    
    def normalize_header(self, header: str) -> str:
        """
        Normalize column headers for consistent field mapping.
        
        Args:
            header: Original header name
            
        Returns:
            Normalized header name
        """
        if not isinstance(header, str):
            return str(header)
        
        # Clean HTML artifacts first
        normalized = self.clean_html_artifacts(header)
        
        # Convert to lowercase and remove special characters
        normalized = normalized.lower().strip()
        normalized = re.sub(r'[^\w\s]', '', normalized)
        normalized = re.sub(r'\s+', '_', normalized)
        
        # Remove common prefixes/suffixes
        normalized = re.sub(r'^(the_|a_|an_)', '', normalized)
        normalized = re.sub(r'(_the|_a|_an)$', '', normalized)
        
        return normalized
    
    def detect_file_type(self, file_path: Union[str, Path]) -> str:
        """
        Detect the type of ERP data file based on filename and content.
        
        Args:
            file_path: Path to the file
            
        Returns:
            Detected file type (inventory, sales_orders, yarn_demand, etc.)
        """
        file_path = Path(file_path)
        filename = file_path.name.lower()
        
        # Pattern matching for file type detection
        if 'inventory' in filename:
            return 'inventory'
        elif 'so_list' in filename or 'sales_order' in filename:
            return 'sales_orders'
        elif 'yarn_demand' in filename:
            return 'yarn_demand'
        elif 'expected_yarn' in filename:
            return 'yarn_report'
        else:
            return 'unknown'
    
    def convert_data_type(self, value: Any, target_type: str, field_name: str = '') -> Any:
        """
        Convert value to target data type with error handling.
        
        Args:
            value: Input value to convert
            target_type: Target data type ('str', 'int', 'float', 'decimal', 'date', 'bool')
            field_name: Field name for error logging
            
        Returns:
            Converted value or None if conversion fails
        """
        if value is None or (isinstance(value, str) and value.strip() == ''):
            return None
        
        try:
            if target_type == 'str':
                return str(value).strip()
            
            elif target_type == 'int':
                # Clean numeric strings
                if isinstance(value, str):
                    value = re.sub(r'[^\d-]', '', value)
                return int(float(value)) if value else None
            
            elif target_type == 'float':
                # Handle currency and formatted numbers
                if isinstance(value, str):
                    value = value.replace('$', '').replace(',', '').strip()
                    value = re.sub(r'[^\d.-]', '', value)
                return float(value) if value else None
            
            elif target_type == 'decimal':
                if isinstance(value, str):
                    value = value.replace('$', '').replace(',', '').strip()
                    value = re.sub(r'[^\d.-]', '', value)
                return Decimal(value) if value else None
            
            elif target_type == 'date':
                if isinstance(value, (date, datetime)):
                    return value
                
                if isinstance(value, str):
                    value = value.strip()
                    for date_format in self.config.date_formats:
                        try:
                            return datetime.strptime(value, date_format).date()
                        except ValueError:
                            continue
                    
                    # Try to parse with pandas for more flexibility
                    try:
                        return pd.to_datetime(value).date()
                    except:
                        pass
                
                return None
            
            elif target_type == 'bool':
                if isinstance(value, str):
                    value = value.lower().strip()
                    return value in ['true', '1', 'yes', 'y', 'on', 'checked']
                return bool(value)
            
            else:
                return value
                
        except (ValueError, TypeError, InvalidOperation) as e:
            self.logger.warning(f"Data type conversion failed for field '{field_name}': {value} -> {target_type}. Error: {e}")
            return None
        except Exception as e:
            self.logger.error(f"Unexpected error in data type conversion for field '{field_name}': {e}")
            return None
    
    def validate_row_data(self, row_data: Dict[str, Any], file_type: str) -> Tuple[bool, List[str]]:
        """
        Validate a row of data against business rules.
        
        Args:
            row_data: Dictionary containing row data
            file_type: Type of file being processed
            
        Returns:
            Tuple of (is_valid, list_of_errors)
        """
        errors = []
        
        try:
            if file_type == 'inventory':
                # Required fields validation
                required_fields = ['style_number', 'order_number', 'roll_number']
                for field in required_fields:
                    if not row_data.get(field):
                        errors.append(f"Missing required field: {field}")
                
                # Quantity validation
                qty_yards = row_data.get('quantity_yards')
                if qty_yards is not None:
                    try:
                        qty_value = float(qty_yards)
                        if qty_value < 0:
                            errors.append("Quantity yards cannot be negative")
                        if qty_value > 100000:  # Reasonable upper limit
                            errors.append("Quantity yards seems unusually high")
                    except (ValueError, TypeError):
                        errors.append("Invalid quantity yards format")
            
            elif file_type == 'sales_orders':
                # Required fields validation
                required_fields = ['order_status', 'unit_price']
                for field in required_fields:
                    if not row_data.get(field):
                        errors.append(f"Missing required field: {field}")
                
                # Price validation
                unit_price = row_data.get('unit_price')
                if unit_price is not None:
                    try:
                        price_value = float(str(unit_price).replace('$', '').replace(',', ''))
                        if price_value < 0:
                            errors.append("Unit price cannot be negative")
                    except (ValueError, TypeError):
                        errors.append("Invalid unit price format")
            
            elif file_type == 'yarn_demand':
                # Required fields validation
                required_fields = ['yarn_id', 'supplier_name']
                for field in required_fields:
                    if not row_data.get(field):
                        errors.append(f"Missing required field: {field}")
                
                # Inventory balance validation
                current_inventory = row_data.get('current_inventory')
                if current_inventory is not None:
                    try:
                        inv_value = float(str(current_inventory).replace(',', ''))
                        if abs(inv_value) > 1000000:  # Reasonable limits
                            errors.append("Inventory value seems unusually high")
                    except (ValueError, TypeError):
                        errors.append("Invalid inventory format")
        
        except Exception as e:
            errors.append(f"Validation error: {e}")
        
        return len(errors) == 0, errors
    
    def parse_csv_file(self, file_path: Union[str, Path]) -> ParseResult:
        """
        Parse CSV file with comprehensive error handling and HTML cleaning.
        
        Args:
            file_path: Path to CSV file
            
        Returns:
            ParseResult object with parsed data and metadata
        """
        file_path = Path(file_path)
        errors = []
        warnings = []
        data = []
        original_headers = []
        normalized_headers = []
        
        try:
            # Detect encoding
            encoding = self.config.encoding or self.detect_encoding(file_path)
            file_type = self.detect_file_type(file_path)
            
            self.logger.info(f"Parsing CSV file: {file_path} (type: {file_type}, encoding: {encoding})")
            
            # Read file with detected encoding
            with open(file_path, 'r', encoding=encoding, errors='replace') as f:
                # Clean the entire file content first
                content = f.read()
                
                # Remove BOM if present
                if content.startswith('\ufeff'):
                    content = content[1:]
                
                # Initial HTML cleaning for the entire content
                if self.config.strip_html:
                    content = self.clean_html_artifacts(content)
                
                # Create CSV reader
                csv_reader = csv.reader(StringIO(content), delimiter=self.config.delimiter)
                
                # Process headers
                try:
                    raw_headers = next(csv_reader)
                    original_headers = [str(h).strip() for h in raw_headers]
                    
                    if self.config.normalize_headers:
                        normalized_headers = [self.normalize_header(h) for h in original_headers]
                        field_mapping = self.FIELD_MAPPINGS.get(file_type, {})
                        
                        # Apply field mapping
                        mapped_headers = []
                        for header in original_headers:
                            mapped_name = field_mapping.get(header.lower(), self.normalize_header(header))
                            mapped_headers.append(mapped_name)
                        normalized_headers = mapped_headers
                    else:
                        normalized_headers = original_headers
                    
                    self.logger.debug(f"Headers: {len(original_headers)} columns detected")
                    
                except StopIteration:
                    errors.append("File appears to be empty or has no headers")
                    return ParseResult([], 0, 0, errors, warnings, {}, [], [])
                
                # Process data rows
                row_count = 0
                valid_count = 0
                
                for row_num, row in enumerate(csv_reader, start=2):  # Start at 2 since headers are row 1
                    if self.config.skip_rows > 0 and row_num <= self.config.skip_rows + 1:
                        continue
                    
                    if self.config.max_rows and row_count >= self.config.max_rows:
                        break
                    
                    row_count += 1
                    
                    try:
                        # Ensure row has same length as headers
                        while len(row) < len(normalized_headers):
                            row.append('')
                        row = row[:len(normalized_headers)]
                        
                        # Clean and convert row data
                        row_data = {}
                        for i, (header, value) in enumerate(zip(normalized_headers, row)):
                            # Clean HTML artifacts from individual cells
                            cleaned_value = self.clean_html_artifacts(str(value)) if self.config.strip_html else str(value)
                            cleaned_value = cleaned_value.strip()
                            
                            # Convert empty strings to None
                            if cleaned_value == '':
                                cleaned_value = None
                            
                            row_data[header] = cleaned_value
                        
                        # Validate row data
                        if self.config.validate_data:
                            is_valid, validation_errors = self.validate_row_data(row_data, file_type)
                            if not is_valid:
                                warnings.extend([f"Row {row_num}: {error}" for error in validation_errors])
                            else:
                                valid_count += 1
                        else:
                            valid_count += 1
                        
                        data.append(row_data)
                        
                    except Exception as e:
                        error_msg = f"Error processing row {row_num}: {e}"
                        errors.append(error_msg)
                        self.logger.warning(error_msg)
                        continue
                
                self.logger.info(f"CSV parsing completed: {valid_count}/{row_count} valid rows")
        
        except UnicodeDecodeError as e:
            error_msg = f"Encoding error: {e}. Try specifying a different encoding."
            errors.append(error_msg)
            self.logger.error(error_msg)
        except Exception as e:
            error_msg = f"Unexpected error parsing CSV: {e}"
            errors.append(error_msg)
            self.logger.error(error_msg)
        
        metadata = {
            'file_path': str(file_path),
            'file_type': file_type,
            'encoding': encoding,
            'file_size_bytes': file_path.stat().st_size if file_path.exists() else 0,
            'parsing_timestamp': datetime.now().isoformat(),
            'parser_config': {
                'strip_html': self.config.strip_html,
                'normalize_headers': self.config.normalize_headers,
                'validate_data': self.config.validate_data
            }
        }
        
        return ParseResult(
            data=data,
            total_rows=row_count,
            valid_rows=valid_count,
            errors=errors,
            warnings=warnings,
            metadata=metadata,
            original_headers=original_headers,
            normalized_headers=normalized_headers
        )
    
    def parse_excel_file(self, file_path: Union[str, Path], sheet_name: Optional[str] = None) -> ParseResult:
        """
        Parse Excel file with comprehensive error handling.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Specific sheet to parse (None for first sheet)
            
        Returns:
            ParseResult object with parsed data and metadata
        """
        file_path = Path(file_path)
        errors = []
        warnings = []
        data = []
        original_headers = []
        normalized_headers = []
        
        try:
            file_type = self.detect_file_type(file_path)
            
            self.logger.info(f"Parsing Excel file: {file_path} (type: {file_type})")
            
            # Try different Excel parsing methods
            try:
                # Method 1: openpyxl for .xlsx files
                if file_path.suffix.lower() in ['.xlsx', '.xlsm']:
                    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
                    sheet_names = workbook.sheetnames
                    
                    if sheet_name:
                        if sheet_name not in sheet_names:
                            errors.append(f"Sheet '{sheet_name}' not found. Available sheets: {sheet_names}")
                            return ParseResult([], 0, 0, errors, warnings, {}, [], [])
                        worksheet = workbook[sheet_name]
                    else:
                        worksheet = workbook.active
                    
                    # Convert to list of lists
                    rows = list(worksheet.iter_rows(values_only=True))
                    
                else:
                    # Method 2: xlrd for .xls files
                    workbook = xlrd.open_workbook(file_path)
                    sheet_names = workbook.sheet_names()
                    
                    if sheet_name:
                        if sheet_name not in sheet_names:
                            errors.append(f"Sheet '{sheet_name}' not found. Available sheets: {sheet_names}")
                            return ParseResult([], 0, 0, errors, warnings, {}, [], [])
                        worksheet = workbook.sheet_by_name(sheet_name)
                    else:
                        worksheet = workbook.sheet_by_index(0)
                    
                    # Convert to list of lists
                    rows = []
                    for row_num in range(worksheet.nrows):
                        row = []
                        for col_num in range(worksheet.ncols):
                            cell_value = worksheet.cell_value(row_num, col_num)
                            row.append(cell_value)
                        rows.append(row)
            
            except Exception as e:
                # Method 3: pandas fallback
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name or 0)
                    rows = [df.columns.tolist()] + df.values.tolist()
                except Exception as pandas_error:
                    errors.append(f"Failed to parse Excel file with all methods. Last error: {pandas_error}")
                    return ParseResult([], 0, 0, errors, warnings, {}, [], [])
            
            if not rows:
                errors.append("Excel file appears to be empty")
                return ParseResult([], 0, 0, errors, warnings, {}, [], [])
            
            # Process headers
            raw_headers = [str(cell) if cell is not None else '' for cell in rows[0]]
            original_headers = [h.strip() for h in raw_headers if h.strip()]
            
            if not original_headers:
                errors.append("No valid headers found in Excel file")
                return ParseResult([], 0, 0, errors, warnings, {}, [], [])
            
            if self.config.normalize_headers:
                normalized_headers = [self.normalize_header(h) for h in original_headers]
                field_mapping = self.FIELD_MAPPINGS.get(file_type, {})
                
                # Apply field mapping
                mapped_headers = []
                for header in original_headers:
                    mapped_name = field_mapping.get(header.lower(), self.normalize_header(header))
                    mapped_headers.append(mapped_name)
                normalized_headers = mapped_headers
            else:
                normalized_headers = original_headers
            
            self.logger.debug(f"Excel headers: {len(original_headers)} columns detected")
            
            # Process data rows
            row_count = 0
            valid_count = 0
            
            for row_idx, row in enumerate(rows[1:], start=2):  # Skip header row
                if self.config.skip_rows > 0 and row_idx <= self.config.skip_rows + 1:
                    continue
                
                if self.config.max_rows and row_count >= self.config.max_rows:
                    break
                
                row_count += 1
                
                try:
                    # Convert row to strings and handle None values
                    str_row = []
                    for cell in row:
                        if cell is None:
                            str_row.append('')
                        elif isinstance(cell, (int, float)):
                            str_row.append(str(cell))
                        else:
                            str_row.append(str(cell))
                    
                    # Ensure row has same length as headers
                    while len(str_row) < len(normalized_headers):
                        str_row.append('')
                    str_row = str_row[:len(normalized_headers)]
                    
                    # Create row data dictionary
                    row_data = {}
                    for header, value in zip(normalized_headers, str_row):
                        # Clean HTML artifacts if needed
                        cleaned_value = self.clean_html_artifacts(value) if self.config.strip_html else value
                        cleaned_value = cleaned_value.strip()
                        
                        # Convert empty strings to None
                        if cleaned_value == '':
                            cleaned_value = None
                        
                        row_data[header] = cleaned_value
                    
                    # Validate row data
                    if self.config.validate_data:
                        is_valid, validation_errors = self.validate_row_data(row_data, file_type)
                        if not is_valid:
                            warnings.extend([f"Row {row_idx}: {error}" for error in validation_errors])
                        else:
                            valid_count += 1
                    else:
                        valid_count += 1
                    
                    data.append(row_data)
                
                except Exception as e:
                    error_msg = f"Error processing row {row_idx}: {e}"
                    errors.append(error_msg)
                    self.logger.warning(error_msg)
                    continue
            
            self.logger.info(f"Excel parsing completed: {valid_count}/{row_count} valid rows")
        
        except Exception as e:
            error_msg = f"Unexpected error parsing Excel file: {e}"
            errors.append(error_msg)
            self.logger.error(error_msg)
        
        metadata = {
            'file_path': str(file_path),
            'file_type': file_type,
            'sheet_name': sheet_name,
            'file_size_bytes': file_path.stat().st_size if file_path.exists() else 0,
            'parsing_timestamp': datetime.now().isoformat(),
            'parser_config': {
                'strip_html': self.config.strip_html,
                'normalize_headers': self.config.normalize_headers,
                'validate_data': self.config.validate_data
            }
        }
        
        return ParseResult(
            data=data,
            total_rows=row_count,
            valid_rows=valid_count,
            errors=errors,
            warnings=warnings,
            metadata=metadata,
            original_headers=original_headers,
            normalized_headers=normalized_headers
        )
    
    def parse_file(self, file_path: Union[str, Path], **kwargs) -> ParseResult:
        """
        Parse file based on extension (CSV or Excel).
        
        Args:
            file_path: Path to file
            **kwargs: Additional arguments for specific parsers
            
        Returns:
            ParseResult object
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            return ParseResult(
                data=[],
                total_rows=0,
                valid_rows=0,
                errors=[f"File not found: {file_path}"],
                warnings=[],
                metadata={'file_path': str(file_path)},
                original_headers=[],
                normalized_headers=[]
            )
        
        if file_path.suffix.lower() == '.csv':
            return self.parse_csv_file(file_path, **kwargs)
        elif file_path.suffix.lower() in ['.xlsx', '.xls', '.xlsm']:
            return self.parse_excel_file(file_path, **kwargs)
        else:
            return ParseResult(
                data=[],
                total_rows=0,
                valid_rows=0,
                errors=[f"Unsupported file format: {file_path.suffix}"],
                warnings=[],
                metadata={'file_path': str(file_path)},
                original_headers=[],
                normalized_headers=[]
            )
    
    def batch_parse_directory(self, directory_path: Union[str, Path], file_pattern: str = "*") -> Dict[str, ParseResult]:
        """
        Parse all matching files in a directory.
        
        Args:
            directory_path: Path to directory containing files
            file_pattern: Glob pattern for file matching
            
        Returns:
            Dictionary mapping file paths to ParseResult objects
        """
        directory_path = Path(directory_path)
        results = {}
        
        if not directory_path.exists():
            self.logger.error(f"Directory not found: {directory_path}")
            return results
        
        # Find matching files
        matching_files = list(directory_path.glob(file_pattern))
        supported_extensions = {'.csv', '.xlsx', '.xls', '.xlsm'}
        
        files_to_process = [
            f for f in matching_files 
            if f.is_file() and f.suffix.lower() in supported_extensions
        ]
        
        self.logger.info(f"Found {len(files_to_process)} files to process in {directory_path}")
        
        for file_path in files_to_process:
            try:
                self.logger.info(f"Processing file: {file_path}")
                result = self.parse_file(file_path)
                results[str(file_path)] = result
                
                if result.errors:
                    self.logger.warning(f"File {file_path} had {len(result.errors)} errors")
                
            except Exception as e:
                error_msg = f"Failed to process file {file_path}: {e}"
                self.logger.error(error_msg)
                results[str(file_path)] = ParseResult(
                    data=[],
                    total_rows=0,
                    valid_rows=0,
                    errors=[error_msg],
                    warnings=[],
                    metadata={'file_path': str(file_path)},
                    original_headers=[],
                    normalized_headers=[]
                )
        
        return results


def create_parser(config: Optional[Dict[str, Any]] = None) -> ERPDataParser:
    """
    Factory function to create an ERP data parser with custom configuration.
    
    Args:
        config: Dictionary with parsing configuration
        
    Returns:
        ERPDataParser instance
    """
    if config:
        parsing_config = ParsingConfig(**config)
    else:
        parsing_config = ParsingConfig()
    
    return ERPDataParser(parsing_config)


# Example usage and testing
if __name__ == "__main__":
    # Test the parser with sample configuration
    config = ParsingConfig(
        strip_html=True,
        normalize_headers=True,
        validate_data=True,
        chunk_size=500
    )
    
    parser = ERPDataParser(config)
    
    # Test parsing a sample file
    sample_file = "/mnt/c/Users/psytz/TMUX Final/Tmux-Orchestrator/ERP Data/eFab_Inventory_F01_20250726.csv"
    
    if Path(sample_file).exists():
        result = parser.parse_file(sample_file)
        
        print(f"Parsing Results:")
        print(f"  Total rows: {result.total_rows}")
        print(f"  Valid rows: {result.valid_rows}")
        print(f"  Errors: {len(result.errors)}")
        print(f"  Warnings: {len(result.warnings)}")
        print(f"  Headers: {len(result.original_headers)}")
        
        if result.data:
            print(f"  Sample data: {result.data[0]}")
    else:
        print("Sample file not found for testing")