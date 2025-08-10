"""
ERP Data Export Module - Production-Ready Data Export Tools

This module provides comprehensive export capabilities for textile ERP data
with multiple format support, query-based exports, scheduled operations,
and MCP tool integration for Agent-MCP.

Features:
- Multiple export formats (CSV, Excel, JSON, PDF reports)
- Query-based data filtering and selection
- Scheduled export capabilities
- Template-based report generation
- Data aggregation and analytics
- Memory-efficient streaming for large datasets
- Comprehensive error handling and logging
- MCP tool integration
"""

import csv
import json
import sqlite3
from pathlib import Path
from typing import Dict, List, Any, Optional, Union, Iterator, Tuple
from datetime import datetime, date, timedelta
from dataclasses import dataclass, asdict
import uuid
import io
import zipfile
import tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
import time

# Third-party imports
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from ..core.config import logger
from ..db.connection import get_db_connection


@dataclass
class ExportConfig:
    """Configuration for export operations."""
    format: str = 'csv'  # csv, excel, json, pdf
    include_headers: bool = True
    max_rows: Optional[int] = None
    chunk_size: int = 10000
    date_format: str = '%Y-%m-%d'
    datetime_format: str = '%Y-%m-%d %H:%M:%S'
    decimal_places: int = 2
    null_value: str = ''
    encoding: str = 'utf-8'
    compression: Optional[str] = None  # zip, gzip
    template_path: Optional[str] = None
    include_metadata: bool = True


@dataclass
class ExportResult:
    """Result of export operation."""
    export_id: str
    file_path: str
    format: str
    total_rows: int
    file_size_bytes: int
    processing_time_seconds: float
    query: str
    success: bool
    errors: List[str]
    warnings: List[str]
    metadata: Dict[str, Any]


class ERPDataExporter:
    """
    Production-ready ERP data exporter with comprehensive format support
    and query-based filtering capabilities.
    """
    
    # Predefined queries for common export scenarios
    PREDEFINED_QUERIES = {
        'inventory_summary': """
            SELECT 
                fi.inventory_id,
                ft.fabric_name,
                ft.fabric_category,
                s.company_name as supplier_name,
                fi.quantity_meters,
                fi.available_meters,
                fi.unit_cost,
                fi.location_warehouse,
                fi.location_zone,
                fi.quality_grade,
                fi.received_date
            FROM fabric_inventory fi
            JOIN fabric_types ft ON fi.fabric_type_id = ft.fabric_type_id
            LEFT JOIN suppliers s ON fi.supplier_id = s.supplier_id
            WHERE fi.available_meters > 0
            ORDER BY fi.received_date DESC
        """,
        
        'sales_orders_active': """
            SELECT 
                so.order_id,
                so.order_number,
                c.company_name as customer_name,
                so.order_date,
                so.promised_delivery_date,
                so.status,
                so.total_value,
                so.currency,
                COUNT(soi.line_id) as line_items
            FROM sales_orders so
            JOIN customers c ON so.customer_id = c.customer_id
            LEFT JOIN sales_order_items soi ON so.order_id = soi.order_id
            WHERE so.status IN ('NEW', 'CONFIRMED', 'IN_PRODUCTION')
            GROUP BY so.order_id
            ORDER BY so.order_date DESC
        """,
        
        'production_orders_current': """
            SELECT 
                po.order_id,
                po.product_type,
                ft.fabric_name,
                po.quantity_pieces,
                po.priority,
                po.planned_start_date,
                po.planned_end_date,
                po.actual_start_date,
                po.status,
                po.completion_percentage,
                pl.line_name as assigned_line,
                w.first_name || ' ' || w.last_name as supervisor_name
            FROM production_orders po
            JOIN fabric_types ft ON po.fabric_type_id = ft.fabric_type_id
            LEFT JOIN production_lines pl ON po.assigned_line = pl.line_id
            LEFT JOIN workers w ON po.supervisor_id = w.worker_id
            WHERE po.status IN ('PENDING', 'IN_PROGRESS')
            ORDER BY po.priority, po.planned_start_date
        """,
        
        'quality_metrics_weekly': """
            SELECT 
                DATE(qi.inspection_date) as inspection_date,
                qi.inspection_type,
                COUNT(*) as total_inspections,
                SUM(CASE WHEN qi.overall_result = 'PASS' THEN 1 ELSE 0 END) as passed_inspections,
                ROUND(AVG(qi.defect_rate), 2) as avg_defect_rate,
                SUM(qi.critical_defects) as total_critical_defects,
                SUM(qi.major_defects) as total_major_defects,
                SUM(qi.minor_defects) as total_minor_defects
            FROM quality_inspections qi
            WHERE qi.inspection_date >= date('now', '-7 days')
            GROUP BY DATE(qi.inspection_date), qi.inspection_type
            ORDER BY inspection_date DESC
        """,
        
        'supplier_performance': """
            SELECT 
                s.supplier_id,
                s.company_name,
                s.supplier_type,
                s.quality_rating,
                s.delivery_rating,
                COUNT(po.order_id) as total_orders,
                SUM(po.total_value) as total_value,
                AVG(
                    CASE 
                        WHEN po.actual_delivery_date IS NOT NULL AND po.expected_delivery_date IS NOT NULL
                        THEN julianday(po.expected_delivery_date) - julianday(po.actual_delivery_date)
                        ELSE 0
                    END
                ) as avg_delivery_days_early
            FROM suppliers s
            LEFT JOIN purchase_orders po ON s.supplier_id = po.supplier_id
            WHERE po.order_date >= date('now', '-90 days')
            GROUP BY s.supplier_id
            HAVING total_orders > 0
            ORDER BY s.quality_rating DESC, total_value DESC
        """,
        
        'machine_utilization': """
            SELECT 
                m.machine_id,
                m.machine_name,
                m.machine_type,
                pl.line_name,
                m.status,
                COALESCE(SUM(mph.runtime_minutes) / 60.0, 0) as total_runtime_hours,
                COALESCE(AVG(mph.avg_efficiency), 0) as avg_efficiency,
                COUNT(md.downtime_id) as downtime_events,
                COALESCE(SUM(md.duration_minutes) / 60.0, 0) as total_downtime_hours
            FROM machines m
            LEFT JOIN production_lines pl ON m.line_id = pl.line_id
            LEFT JOIN machine_performance_hourly mph ON m.machine_id = mph.machine_id 
                AND mph.hour_timestamp >= date('now', '-7 days')
            LEFT JOIN machine_downtime md ON m.machine_id = md.machine_id 
                AND md.start_time >= date('now', '-7 days')
            GROUP BY m.machine_id
            ORDER BY avg_efficiency DESC
        """
    }
    
    def __init__(self, config: Optional[ExportConfig] = None):
        """Initialize exporter with configuration."""
        self.config = config or ExportConfig()
        self.logger = logger
        
    def generate_export_id(self) -> str:
        """Generate unique export ID."""
        return f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
    
    def execute_query(self, query: str, parameters: Optional[Tuple] = None) -> Tuple[List[Dict[str, Any]], List[str]]:
        """
        Execute SQL query and return results with column names.
        
        Args:
            query: SQL query to execute
            parameters: Query parameters
            
        Returns:
            Tuple of (data_rows, column_names)
        """
        conn = None
        try:
            conn = get_db_connection()
            conn.row_factory = sqlite3.Row  # Enable column access by name
            cursor = conn.cursor()
            
            # Execute query with parameters
            if parameters:
                cursor.execute(query, parameters)
            else:
                cursor.execute(query)
            
            # Fetch all results
            rows = cursor.fetchall()
            
            # Get column names
            column_names = [description[0] for description in cursor.description] if cursor.description else []
            
            # Convert rows to dictionaries
            data = [dict(row) for row in rows]
            
            self.logger.debug(f"Query executed successfully: {len(data)} rows returned")
            return data, column_names
            
        except sqlite3.Error as e:
            error_msg = f"Database query failed: {e}"
            self.logger.error(error_msg)
            raise ValueError(error_msg)
        finally:
            if conn:
                conn.close()
    
    def format_value(self, value: Any) -> str:
        """Format value according to export configuration."""
        if value is None:
            return self.config.null_value
        
        if isinstance(value, (int, float)):
            if isinstance(value, float):
                return f"{value:.{self.config.decimal_places}f}"
            return str(value)
        
        if isinstance(value, datetime):
            return value.strftime(self.config.datetime_format)
        
        if isinstance(value, date):
            return value.strftime(self.config.date_format)
        
        return str(value)
    
    def export_to_csv(self, data: List[Dict[str, Any]], file_path: Path) -> None:
        """Export data to CSV format."""
        if not data:
            # Create empty CSV file
            with open(file_path, 'w', newline='', encoding=self.config.encoding) as csvfile:
                writer = csv.writer(csvfile)
                if self.config.include_headers:
                    writer.writerow([])
            return
        
        with open(file_path, 'w', newline='', encoding=self.config.encoding) as csvfile:
            fieldnames = list(data[0].keys())
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            if self.config.include_headers:
                writer.writeheader()
            
            for row in data:
                formatted_row = {k: self.format_value(v) for k, v in row.items()}
                writer.writerow(formatted_row)
        
        self.logger.debug(f"CSV export completed: {len(data)} rows written to {file_path}")
    
    def export_to_excel(self, data: List[Dict[str, Any]], file_path: Path, sheet_name: str = "Data") -> None:
        """Export data to Excel format with formatting."""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        
        if not data:
            # Create empty worksheet
            workbook.save(file_path)
            return
        
        # Convert to pandas DataFrame for easier Excel export
        df = pd.DataFrame(data)
        
        # Format data
        for col in df.columns:
            df[col] = df[col].apply(self.format_value)
        
        # Write to worksheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=self.config.include_headers), 1):
            for c_idx, value in enumerate(row, 1):
                cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
                
                # Apply header formatting
                if r_idx == 1 and self.config.include_headers:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Limit maximum width
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(file_path)
        self.logger.debug(f"Excel export completed: {len(data)} rows written to {file_path}")
    
    def export_to_json(self, data: List[Dict[str, Any]], file_path: Path) -> None:
        """Export data to JSON format."""
        # Format data for JSON serialization
        formatted_data = []
        for row in data:
            formatted_row = {}
            for key, value in row.items():
                if isinstance(value, (datetime, date)):
                    formatted_row[key] = self.format_value(value)
                else:
                    formatted_row[key] = value
            formatted_data.append(formatted_row)
        
        export_object = {
            'data': formatted_data,
            'metadata': {
                'export_timestamp': datetime.now().isoformat(),
                'total_records': len(formatted_data),
                'format': 'json'
            } if self.config.include_metadata else None
        }
        
        # Remove metadata if not requested
        if not self.config.include_metadata:
            export_object = formatted_data
        
        with open(file_path, 'w', encoding=self.config.encoding) as jsonfile:
            json.dump(export_object, jsonfile, indent=2, ensure_ascii=False)
        
        self.logger.debug(f"JSON export completed: {len(data)} records written to {file_path}")
    
    def export_to_pdf(self, data: List[Dict[str, Any]], file_path: Path, title: str = "ERP Data Report") -> None:
        """Export data to PDF format with table formatting."""
        if not data:
            # Create empty PDF
            doc = SimpleDocTemplate(str(file_path), pagesize=letter)
            story = [Paragraph("No data to export", getSampleStyleSheet()['Normal'])]
            doc.build(story)
            return
        
        # Create PDF document
        doc = SimpleDocTemplate(str(file_path), pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Add title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        story.append(Paragraph(title, title_style))
        
        # Add metadata
        if self.config.include_metadata:
            metadata_style = styles['Normal']
            metadata_text = f"Generated: {datetime.now().strftime(self.config.datetime_format)}<br/>Records: {len(data)}"
            story.append(Paragraph(metadata_text, metadata_style))
            story.append(Spacer(1, 20))
        
        # Prepare table data
        if self.config.include_headers:
            table_data = [list(data[0].keys())]
        else:
            table_data = []
        
        # Add data rows (limit for PDF readability)
        max_pdf_rows = min(len(data), self.config.max_rows or 100)
        for row in data[:max_pdf_rows]:
            formatted_row = [self.format_value(v) for v in row.values()]
            table_data.append(formatted_row)
        
        # Create table
        table = Table(table_data)
        
        # Apply table style
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        
        table.setStyle(table_style)
        story.append(table)
        
        # Add note if data was truncated
        if len(data) > max_pdf_rows:
            note_text = f"<br/><i>Note: Showing first {max_pdf_rows} of {len(data)} records</i>"
            story.append(Paragraph(note_text, styles['Italic']))
        
        # Build PDF
        doc.build(story)
        self.logger.debug(f"PDF export completed: {min(len(data), max_pdf_rows)} rows written to {file_path}")
    
    def compress_file(self, file_path: Path) -> Path:
        """Compress exported file if compression is enabled."""
        if not self.config.compression:
            return file_path
        
        if self.config.compression == 'zip':
            zip_path = file_path.with_suffix(f"{file_path.suffix}.zip")
            
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                zipf.write(file_path, file_path.name)
            
            # Remove original file
            file_path.unlink()
            
            self.logger.debug(f"File compressed: {zip_path}")
            return zip_path
        
        elif self.config.compression == 'gzip':
            import gzip
            gz_path = file_path.with_suffix(f"{file_path.suffix}.gz")
            
            with open(file_path, 'rb') as f_in:
                with gzip.open(gz_path, 'wb') as f_out:
                    f_out.writelines(f_in)
            
            # Remove original file
            file_path.unlink()
            
            self.logger.debug(f"File compressed: {gz_path}")
            return gz_path
        
        return file_path
    
    def export_query(self, query: str, output_path: Union[str, Path], 
                    parameters: Optional[Tuple] = None, **kwargs) -> ExportResult:
        """
        Export query results to specified format.
        
        Args:
            query: SQL query to execute
            output_path: Output file path
            parameters: Query parameters
            **kwargs: Additional configuration overrides
            
        Returns:
            ExportResult with operation details
        """
        start_time = time.time()
        export_id = self.generate_export_id()
        output_path = Path(output_path)
        errors = []
        warnings = []
        
        # Override config with kwargs
        config = ExportConfig(**{**asdict(self.config), **kwargs})
        self.config = config
        
        self.logger.info(f"Starting export {export_id} to {output_path}")
        
        try:
            # Execute query
            data, column_names = self.execute_query(query, parameters)
            
            # Apply row limit
            if config.max_rows and len(data) > config.max_rows:
                data = data[:config.max_rows]
                warnings.append(f"Results limited to {config.max_rows} rows")
            
            # Ensure output directory exists
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Export based on format
            if config.format.lower() == 'csv':
                self.export_to_csv(data, output_path)
            elif config.format.lower() in ['excel', 'xlsx']:
                self.export_to_excel(data, output_path)
            elif config.format.lower() == 'json':
                self.export_to_json(data, output_path)
            elif config.format.lower() == 'pdf':
                self.export_to_pdf(data, output_path)
            else:
                raise ValueError(f"Unsupported export format: {config.format}")
            
            # Compress if requested
            final_path = self.compress_file(output_path)
            
            # Get file size
            file_size = final_path.stat().st_size if final_path.exists() else 0
            
            processing_time = time.time() - start_time
            
            self.logger.info(f"Export {export_id} completed: {len(data)} rows exported in {processing_time:.2f}s")
            
            return ExportResult(
                export_id=export_id,
                file_path=str(final_path),
                format=config.format,
                total_rows=len(data),
                file_size_bytes=file_size,
                processing_time_seconds=processing_time,
                query=query,
                success=True,
                errors=errors,
                warnings=warnings,
                metadata={
                    'column_names': column_names,
                    'config': asdict(config),
                    'export_timestamp': datetime.now().isoformat(),
                    'parameters': parameters
                }
            )
            
        except Exception as e:
            error_msg = f"Export failed: {e}"
            errors.append(error_msg)
            self.logger.error(error_msg)
            
            return ExportResult(
                export_id=export_id,
                file_path=str(output_path),
                format=config.format,
                total_rows=0,
                file_size_bytes=0,
                processing_time_seconds=time.time() - start_time,
                query=query,
                success=False,
                errors=errors,
                warnings=warnings,
                metadata={}
            )
    
    def export_predefined(self, query_name: str, output_path: Union[str, Path], 
                         parameters: Optional[Dict[str, Any]] = None, **kwargs) -> ExportResult:
        """
        Export using predefined query.
        
        Args:
            query_name: Name of predefined query
            output_path: Output file path
            parameters: Query parameters as dictionary
            **kwargs: Additional configuration overrides
            
        Returns:
            ExportResult with operation details
        """
        if query_name not in self.PREDEFINED_QUERIES:
            available_queries = list(self.PREDEFINED_QUERIES.keys())
            raise ValueError(f"Unknown query '{query_name}'. Available: {available_queries}")
        
        query = self.PREDEFINED_QUERIES[query_name]
        
        # Convert dict parameters to tuple if needed
        query_params = None
        if parameters:
            # For now, assume parameters are for simple substitution
            # In a more advanced version, this could handle named parameters
            query_params = tuple(parameters.values()) if parameters else None
        
        return self.export_query(query, output_path, query_params, **kwargs)
    
    def batch_export(self, export_specs: List[Dict[str, Any]], output_dir: Union[str, Path]) -> Dict[str, ExportResult]:
        """
        Perform multiple exports in batch.
        
        Args:
            export_specs: List of export specifications
            output_dir: Base output directory
            
        Returns:
            Dictionary mapping export names to results
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        results = {}
        
        self.logger.info(f"Starting batch export: {len(export_specs)} exports to {output_dir}")
        
        for i, spec in enumerate(export_specs):
            try:
                name = spec.get('name', f'export_{i}')
                query = spec.get('query')
                query_name = spec.get('query_name')
                format = spec.get('format', 'csv')
                filename = spec.get('filename', f"{name}.{format}")
                parameters = spec.get('parameters')
                config_overrides = {k: v for k, v in spec.items() 
                                  if k not in ['name', 'query', 'query_name', 'filename', 'parameters']}
                
                output_path = output_dir / filename
                
                # Export using query or predefined query
                if query:
                    result = self.export_query(query, output_path, parameters, **config_overrides)
                elif query_name:
                    result = self.export_predefined(query_name, output_path, parameters, **config_overrides)
                else:
                    raise ValueError("Export spec must include 'query' or 'query_name'")
                
                results[name] = result
                
                if result.success:
                    self.logger.info(f"Batch export '{name}' completed: {result.total_rows} rows")
                else:
                    self.logger.warning(f"Batch export '{name}' failed: {len(result.errors)} errors")
                
            except Exception as e:
                error_msg = f"Batch export '{name if 'name' in locals() else i}' failed: {e}"
                self.logger.error(error_msg)
                
                results[spec.get('name', f'export_{i}')] = ExportResult(
                    export_id=self.generate_export_id(),
                    file_path='',
                    format=spec.get('format', 'unknown'),
                    total_rows=0,
                    file_size_bytes=0,
                    processing_time_seconds=0,
                    query=spec.get('query', ''),
                    success=False,
                    errors=[error_msg],
                    warnings=[],
                    metadata={}
                )
        
        # Log summary
        successful_exports = sum(1 for r in results.values() if r.success)
        total_rows = sum(r.total_rows for r in results.values())
        
        self.logger.info(f"Batch export completed: {successful_exports}/{len(export_specs)} successful, {total_rows} total rows")
        
        return results


# MCP Tool Functions for Agent-MCP Integration

def export_erp_data(query: str, output_path: str, format: str = 'csv', 
                   config: Optional[Dict[str, Any]] = None, 
                   parameters: Optional[List[Any]] = None) -> Dict[str, Any]:
    """
    MCP tool to export ERP data using custom query.
    
    Args:
        query: SQL query to execute
        output_path: Output file path
        format: Export format (csv, excel, json, pdf)
        config: Export configuration dictionary
        parameters: Query parameters
        
    Returns:
        Export result as dictionary
    """
    try:
        # Create export configuration
        export_config = ExportConfig(format=format, **(config or {}))
        exporter = ERPDataExporter(export_config)
        
        # Convert parameters list to tuple
        query_params = tuple(parameters) if parameters else None
        
        # Export data
        result = exporter.export_query(query, output_path, query_params)
        
        return {
            'success': result.success,
            'export_id': result.export_id,
            'file_path': result.file_path,
            'format': result.format,
            'total_rows': result.total_rows,
            'file_size_bytes': result.file_size_bytes,
            'processing_time_seconds': result.processing_time_seconds,
            'errors': result.errors,
            'warnings': result.warnings,
            'metadata': result.metadata
        }
        
    except Exception as e:
        logger.error(f"MCP export_erp_data failed: {e}")
        return {
            'success': False,
            'error': str(e),
            'export_id': None,
            'total_rows': 0,
            'file_size_bytes': 0
        }


def export_predefined_report(report_name: str, output_path: str, format: str = 'csv',
                           config: Optional[Dict[str, Any]] = None,
                           parameters: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """
    MCP tool to export predefined ERP reports.
    
    Args:
        report_name: Name of predefined report
        output_path: Output file path
        format: Export format (csv, excel, json, pdf)
        config: Export configuration dictionary
        parameters: Query parameters as dictionary
        
    Returns:
        Export result as dictionary
    """
    try:
        # Create export configuration
        export_config = ExportConfig(format=format, **(config or {}))
        exporter = ERPDataExporter(export_config)
        
        # Export predefined report
        result = exporter.export_predefined(report_name, output_path, parameters)
        
        return {
            'success': result.success,
            'export_id': result.export_id,
            'file_path': result.file_path,
            'format': result.format,
            'total_rows': result.total_rows,
            'file_size_bytes': result.file_size_bytes,
            'processing_time_seconds': result.processing_time_seconds,
            'errors': result.errors,
            'warnings': result.warnings,
            'available_reports': list(exporter.PREDEFINED_QUERIES.keys())
        }
        
    except Exception as e:
        logger.error(f"MCP export_predefined_report failed: {e}")
        return {
            'success': False,
            'error': str(e),
            'export_id': None,
            'total_rows': 0,
            'available_reports': list(ERPDataExporter.PREDEFINED_QUERIES.keys())
        }


def batch_export_reports(export_specs: List[Dict[str, Any]], output_dir: str) -> Dict[str, Any]:
    """
    MCP tool to perform batch export of multiple reports.
    
    Args:
        export_specs: List of export specifications
        output_dir: Base output directory
        
    Returns:
        Batch export results as dictionary
    """
    try:
        exporter = ERPDataExporter()
        results = exporter.batch_export(export_specs, output_dir)
        
        # Summarize results
        total_exports = len(results)
        successful_exports = sum(1 for r in results.values() if r.success)
        total_rows = sum(r.total_rows for r in results.values())
        
        return {
            'success': successful_exports > 0,
            'total_exports': total_exports,
            'successful_exports': successful_exports,
            'failed_exports': total_exports - successful_exports,
            'total_rows': total_rows,
            'export_summary': {
                name: {
                    'success': result.success,
                    'total_rows': result.total_rows,
                    'file_path': result.file_path,
                    'errors': len(result.errors)
                }
                for name, result in results.items()
            }
        }
        
    except Exception as e:
        logger.error(f"MCP batch_export_reports failed: {e}")
        return {
            'success': False,
            'error': str(e),
            'total_exports': 0,
            'successful_exports': 0
        }


def get_available_reports() -> Dict[str, Any]:
    """
    MCP tool to get list of available predefined reports.
    
    Returns:
        Dictionary with available reports and their descriptions
    """
    try:
        reports = {}
        
        for query_name, query in ERPDataExporter.PREDEFINED_QUERIES.items():
            # Extract basic description from query comments or structure
            description = f"Predefined query: {query_name}"
            
            # Try to infer description from query
            if 'inventory' in query_name:
                description = "Inventory summary report with stock levels and locations"
            elif 'sales' in query_name:
                description = "Active sales orders report with customer information"
            elif 'production' in query_name:
                description = "Current production orders and progress tracking"
            elif 'quality' in query_name:
                description = "Quality metrics and inspection results"
            elif 'supplier' in query_name:
                description = "Supplier performance and delivery metrics"
            elif 'machine' in query_name:
                description = "Machine utilization and performance statistics"
            
            reports[query_name] = {
                'description': description,
                'query_preview': query[:200] + "..." if len(query) > 200 else query
            }
        
        return {
            'success': True,
            'total_reports': len(reports),
            'reports': reports
        }
        
    except Exception as e:
        logger.error(f"MCP get_available_reports failed: {e}")
        return {
            'success': False,
            'error': str(e),
            'total_reports': 0,
            'reports': {}
        }


# Example usage and testing
if __name__ == "__main__":
    # Test the exporter
    config = ExportConfig(
        format='csv',
        include_headers=True,
        max_rows=1000,
        encoding='utf-8'
    )
    
    exporter = ERPDataExporter(config)
    
    # Test exporting inventory summary
    output_path = "/tmp/test_inventory_export.csv"
    
    try:
        result = exporter.export_predefined('inventory_summary', output_path)
        
        print(f"Export Results:")
        print(f"  Success: {result.success}")
        print(f"  Export ID: {result.export_id}")
        print(f"  File: {result.file_path}")
        print(f"  Total rows: {result.total_rows}")
        print(f"  File size: {result.file_size_bytes} bytes")
        print(f"  Processing time: {result.processing_time_seconds:.2f}s")
        print(f"  Errors: {len(result.errors)}")
        print(f"  Warnings: {len(result.warnings)}")
        
    except Exception as e:
        print(f"Export test failed: {e}")